from openpyxl import Workbook

from .profit import Profit

"""
Клаcc для записи расчетов
пока только в xlsx с помощью openpyxl https://youtu.be/dn3Oi7oaMT4
"""
class Write:

    def __init__(self):
        self.profit = Profit()
        self.api = self.profit.api

    def write_xlsx(self):
        """
        Пишу операции и виртуальные продажи в XLSX файл
        :return:
        """
        filename = "data.xlsx"
        wb = Workbook()
        wb.remove(wb.active)

        # ТАБЛИЦА ОПЕРАЦИЙ с фильтрами
        sheet_operations = wb.create_sheet("Операции")

        # заголовки столбцов
        headers = ['time', 'operation.type', 'value', 'currency', 'instrument.type', 'tiker', 'count', 'is_virtual', 'expected']
        sheet_operations.append(headers)
        # автофильтр
        sheet_operations.auto_filter.ref = "A1:H9999"

        # форматирование заголовка
        for i in range(1, sheet_operations.max_column+1):
            cell = sheet_operations.cell(1, i)
            cell.style = "Good"
            # ширина столбца
            sheet_operations.column_dimensions[cell.column_letter].width = max([len(str(s[i-1]))+1 for s in self.api.data if len(s) >= i])

        # наполнение данными
        for r in self.api.data: sheet_operations.append(r)

        # расчеты по фильтрам
        sheet_operations["J1"].value = "=SUBTOTAL(9,C2:C9999)" # сумма отфильтрованных
        sheet_operations["K1"].value = "=SUBTOTAL(3,C2:C9999)" # колво
        sheet_operations["L1"].value = "=SUBTOTAL(1,C2:C9999)" # среднее
        sheet_operations["M1"].value = self.profit.usdrur # курс для ручной сверки

        # СВОДНАЯ ТАБЛИЦА доходности
        sheet_profit = wb.create_sheet("Доходность")

        sheet_profit.column_dimensions["A"].width = 30

        sheet_profit.append(["Общий доход", self.profit.total, "руб."])
        sheet_profit.append(["Завел денег", self.profit.payed_in, "руб."])
        sheet_profit.append(["Период инвестирования", self.profit.period_days, "дн."])
        sheet_profit.append(["Ориентировочная доходность", self.profit.year_proc, "% год"])

        sheet_profit.append([""])
        sheet_profit.append(["Налогов", self.profit.taxes, "руб."])
        sheet_profit.append(["", self.profit.taxes_proc, "% от заведенных"])
        sheet_profit.append([""])
        sheet_profit.append(["Комиссии", self.profit.comissions, "руб."])
        sheet_profit.append(["", self.profit.comissions_proc, "% от заведенных"])


        wb.active = sheet_profit
        wb.save(filename)